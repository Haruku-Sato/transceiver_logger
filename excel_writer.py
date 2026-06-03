import re
import xlwings as xw
from datetime import datetime

_target_path: str | None = None
_target_sheet: str | None = None
_last_row: int | None = None

HEADERS = ["時刻", "話者", "発言内容", "タグ", "要約"]

# 改行の基準
_BREAK_MIN = 8   # 現在の行がこれ以上のとき
_BREAK_MAX = 10  # 次を足すとこれ以上になるなら改行


def _add_linebreaks(text: str) -> str:
    """句読点を基準に適切な位置で改行を挿入する。

    ルール：
    - 現在の蓄積が _BREAK_MIN 文字以上
    - かつ次のセグメントを足すと _BREAK_MAX 文字以上になる
    → 改行を挿入
    """
    # 句読点の直後で分割（句読点はセグメントに含める）
    segments = [s for s in re.split(r'(?<=[。、！？,.!?])', text) if s]
    if not segments:
        return text

    lines: list[str] = []
    current = ""

    for seg in segments:
        if not current:
            current = seg
            continue
        if len(current) >= _BREAK_MIN and len(current) + len(seg) >= _BREAK_MAX:
            lines.append(current)
            current = seg
        else:
            current += seg

    if current:
        lines.append(current)

    return "\n".join(lines)


def set_target(path: str) -> None:
    global _target_path, _target_sheet, _last_row
    _target_path = path
    _target_sheet = None
    _last_row = None


def set_sheet(name: str | None) -> None:
    global _target_sheet, _last_row
    _target_sheet = name
    _last_row = None


def get_target() -> str | None:
    return _target_path


def get_sheets(path: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _get_ws(wb):
    if _target_sheet:
        return wb.sheets[_target_sheet]
    return wb.sheets[0]


def _next_row(ws) -> int:
    if ws.range("A1").value is None:
        return 1
    row = 1
    while ws.range(f"A{row}").value is not None:
        row += 1
    return row


def _apply_format(ws, row: int) -> None:
    """折り返し設定 + 列幅・行高さの自動調整"""
    cell = ws.range(f"C{row}")
    cell.wrap_text = True
    ws.used_range.columns.autofit()
    ws.used_range.rows.autofit()


def append_row(speaker: str, text: str, tag: str = "", summary: str = "") -> int:
    global _last_row

    if not _target_path:
        raise ValueError("出力先Excelファイルが設定されていません")

    wb = xw.Book(_target_path)
    ws = _get_ws(wb)

    next_row = _next_row(ws)
    if next_row == 1:
        ws.range("A1").value = HEADERS
        next_row = 2

    timestamp = datetime.now().strftime("%H:%M:%S")
    formatted_text = _add_linebreaks(text)
    ws.range(f"A{next_row}").value = [timestamp, speaker, formatted_text, tag, summary]
    _apply_format(ws, next_row)

    _last_row = next_row
    return next_row


def update_last_row(speaker: str, text: str, tag: str = "", summary: str = "") -> None:
    if not _target_path:
        raise ValueError("出力先Excelファイルが設定されていません")
    if _last_row is None:
        raise ValueError("修正できる直前の記録がありません")

    wb = xw.Book(_target_path)
    ws = _get_ws(wb)

    existing_time = ws.range(f"A{_last_row}").value or datetime.now().strftime("%H:%M:%S")
    formatted_text = _add_linebreaks(text)
    ws.range(f"A{_last_row}").value = [existing_time, speaker, formatted_text, tag, summary]
    _apply_format(ws, _last_row)
