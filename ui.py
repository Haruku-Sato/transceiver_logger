import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import queue
import concurrent.futures
import time
import math
import numpy as np
from datetime import datetime

import audio_capture as ac
import vad as vad_module
import speaker_id
import db_manager
import excel_writer
import vocab_manager
import config_manager
import ai_client as ai_module
import supabase_client
from tksheet import Sheet as TkSheet

SAMPLE_RATE = 16000

# 信号レベルメーターのバー数
METER_BARS = 12


class App:
    def __init__(self, root: tk.Tk, mode: str = "offline"):
        self.root = root
        self._mode = mode  # "online" or "offline"
        self.root.title("トランシーバ記録システム")
        self.root.geometry("1280x720")
        self.root.minsize(800, 500)
        self.root.resizable(True, True)

        self.is_recording = False
        self.is_db_recording = False
        self._result_queue: queue.Queue = queue.Queue()
        self._utterance_queue: queue.Queue = queue.Queue()

        self._pending_speaker = "不明"
        self._pending_timestamp = ""

        self._current_samples: list[tuple[np.ndarray, str]] = []
        self._db_capturer: ac.AudioCapture | None = None
        self._db_audio_buf: list[np.ndarray] = []

        self._capturer: ac.AudioCapture | None = None
        self._detector: vad_module.PTTDetector | vad_module.VAD | None = None
        self._vad_thread: threading.Thread | None = None

        self._device_options: list[tuple[int | None, str]] = []
        self._device_var = tk.StringVar()
        self._detect_mode = tk.StringVar(value="ptt")   # "ptt" or "vad"
        self._ptt_threshold_var = tk.StringVar(value="0.002")

        self._known_sheets: list[str] = []
        self._last_sheet_check: float = 0.0

        # AI関連
        self._ai_cfg: dict = config_manager.load_config()
        self._ai_client: ai_module.AIClient | None = None
        self._ai_context: list[dict] = []   # 直近発話コンテキスト
        self._session_cost_jpy: float = 0.0
        self._budget_exhausted: bool = False
        self._budget_warned: bool = False
        self._batch_mode: bool = False       # Trueのとき発話ごとAI呼び出しをスキップ
        self._batch_pending: list[dict] = [] # バッチ処理待ち発話

        # Supabase
        self._session_id: str = ""
        # オンラインモードで起動した場合はweb_modeを強制ON
        default_web = True if mode == "online" else self._ai_cfg.get("web_mode", False)
        self._web_mode_var = tk.BooleanVar(value=default_web)

        # シートビューア
        self._viewer_sheet: TkSheet | None = None
        self._last_viewer_refresh: float = 0.0
        self._viewer_row_count: int = 0      # 前回の行数（差分更新用）

        self._build_device_selector()
        self._build_ui()
        self._start_process_worker()
        self._poll_queue()
        self._apply_ai_config()
        self._apply_supabase_config()

    # ─────────────────────────────── UI構築

    def _build_device_selector(self):
        import sounddevice as sd

        row = tk.Frame(self.root)
        row.pack(fill="x", padx=10, pady=(6, 0))
        tk.Label(row, text="入力デバイス：").pack(side="left")

        self._device_options = [(None, "デフォルト（システム設定に従う）")]
        for i, d in enumerate(sd.query_devices()):
            if d["max_input_channels"] > 0:
                self._device_options.append((i, f"{i}: {d['name']}"))

        labels = [lbl for _, lbl in self._device_options]
        self._device_var.set(labels[0])

        ttk.Combobox(
            row, textvariable=self._device_var,
            values=labels, state="readonly", width=42
        ).pack(side="left", padx=6)

    def _get_selected_device(self) -> int | None:
        label = self._device_var.get()
        for idx, lbl in self._device_options:
            if lbl == label:
                return idx
        return None

    def _build_ui(self):
        paned = tk.PanedWindow(self.root, orient="horizontal", sashrelief="groove", sashwidth=6)
        paned.pack(fill="both", expand=True, padx=4, pady=4)

        # ── 左ペイン：既存タブ
        left_frame = tk.Frame(paned)
        paned.add(left_frame, minsize=480, width=580)

        notebook = ttk.Notebook(left_frame)
        notebook.pack(fill="both", expand=True)

        self._tab_db = ttk.Frame(notebook)
        notebook.add(self._tab_db, text="DB作成")
        self._build_db_tab()

        self._tab_rec = ttk.Frame(notebook)
        notebook.add(self._tab_rec, text="記録・出力")
        self._build_rec_tab()

        self._tab_vocab = ttk.Frame(notebook)
        notebook.add(self._tab_vocab, text="語彙設定")
        self._build_vocab_tab()

        self._tab_ai = ttk.Frame(notebook)
        notebook.add(self._tab_ai, text="AI設定")
        self._build_ai_tab()

        self._tab_cloud = ttk.Frame(notebook)
        notebook.add(self._tab_cloud, text="クラウド設定")
        self._build_cloud_tab()

        # ── 右ペイン：Excelビューア
        right_frame = tk.Frame(paned)
        paned.add(right_frame, minsize=300)
        self._build_sheet_viewer(right_frame)

    def _build_db_tab(self):
        f = self._tab_db

        row = tk.Frame(f)
        row.pack(fill="x", padx=10, pady=(12, 4))
        tk.Label(row, text="名前：").pack(side="left")
        self._db_name_var = tk.StringVar()
        tk.Entry(row, textvariable=self._db_name_var, width=28).pack(side="left", padx=6)

        row2 = tk.Frame(f)
        row2.pack(fill="x", padx=10, pady=4)
        self._btn_db_start = tk.Button(
            row2, text="録音開始", width=10, bg="#d2e6ff",
            command=self._db_start_recording
        )
        self._btn_db_start.pack(side="left", padx=4)
        self._btn_db_stop = tk.Button(
            row2, text="録音停止", width=10, bg="#e1e1e1",
            state="disabled", command=self._db_stop_recording
        )
        self._btn_db_stop.pack(side="left", padx=4)

        tk.Label(f, text="登録済みサンプル：", anchor="w").pack(fill="x", padx=10, pady=(6, 2))

        list_frame = tk.Frame(f)
        list_frame.pack(fill="both", expand=True, padx=10)
        sb = tk.Scrollbar(list_frame, orient="vertical")
        self._sample_lb = tk.Listbox(list_frame, height=7, yscrollcommand=sb.set)
        sb.config(command=self._sample_lb.yview)
        self._sample_lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        tk.Button(f, text="削除", command=self._db_delete_sample).pack(anchor="e", padx=10)
        tk.Button(
            f, text="DBに保存", width=14, bg="#c8f0d8",
            font=("", 11, "bold"), command=self._db_save
        ).pack(pady=10)
        tk.Label(f, text="※ 名前ごとに手順を繰り返す", fg="gray").pack()

    def _build_vocab_tab(self):
        f = self._tab_vocab

        tk.Label(f, text="Whisperに渡す頻出語彙（誤変換抑制用）", anchor="w", fg="gray").pack(
            fill="x", padx=10, pady=(10, 2)
        )

        list_frame = tk.Frame(f)
        list_frame.pack(fill="both", expand=True, padx=10)
        sb = tk.Scrollbar(list_frame, orient="vertical")
        self._vocab_lb = tk.Listbox(list_frame, height=12, yscrollcommand=sb.set)
        sb.config(command=self._vocab_lb.yview)
        self._vocab_lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        add_row = tk.Frame(f)
        add_row.pack(fill="x", padx=10, pady=(6, 2))
        self._vocab_entry = tk.Entry(add_row, width=24)
        self._vocab_entry.pack(side="left", padx=(0, 6))
        self._vocab_entry.bind("<Return>", lambda e: self._vocab_add())
        tk.Button(add_row, text="追加", width=6, command=self._vocab_add).pack(side="left")
        tk.Button(add_row, text="削除", width=6, command=self._vocab_delete).pack(side="left", padx=6)

        self._vocab_load()

    def _vocab_load(self):
        """vocab_manager からリストを読み込んで Listbox に反映する"""
        words = vocab_manager.load_vocab()
        self._vocab_lb.delete(0, "end")
        for w in words:
            self._vocab_lb.insert("end", w)

    def _vocab_add(self):
        word = self._vocab_entry.get().strip()
        if not word:
            return
        existing = list(self._vocab_lb.get(0, "end"))
        if word in existing:
            self._vocab_entry.delete(0, "end")
            return
        self._vocab_lb.insert("end", word)
        self._vocab_entry.delete(0, "end")
        self._vocab_save()

    def _vocab_delete(self):
        sel = self._vocab_lb.curselection()
        if not sel:
            return
        self._vocab_lb.delete(sel[0])
        self._vocab_save()

    def _vocab_save(self):
        words = list(self._vocab_lb.get(0, "end"))
        vocab_manager.save_vocab(words)

    # ─────────────────────────────── クラウド設定タブ

    def _build_cloud_tab(self):
        f = self._tab_cloud
        cfg = self._ai_cfg

        tk.Label(f, text="Supabase設定", font=("", 11, "bold"), anchor="w").pack(
            fill="x", padx=10, pady=(12, 4)
        )
        tk.Label(
            f,
            text="発話ログをSupabaseに送信し、ブラウザから閲覧できます。",
            fg="gray", anchor="w", wraplength=440, justify="left"
        ).pack(fill="x", padx=10, pady=(0, 8))

        # Webモード
        tk.Label(f, text="動作モード", font=("", 10, "bold"), anchor="w").pack(
            fill="x", padx=10, pady=(4, 2)
        )
        mode_frame = tk.Frame(f)
        mode_frame.pack(fill="x", padx=10, pady=(0, 6))
        tk.Radiobutton(
            mode_frame, text="ローカルモード（文字起こし・AI要約をPC上で実行）",
            variable=self._web_mode_var, value=False
        ).pack(anchor="w")
        tk.Radiobutton(
            mode_frame, text="Webモード（音声をクラウドに送信、ブラウザ側で解析）",
            variable=self._web_mode_var, value=True
        ).pack(anchor="w")
        tk.Label(
            f, text="※ Webモードはローカルモードより遅延が発生します",
            fg="gray", font=("", 9)
        ).pack(anchor="w", padx=10)

        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=6)

        self._sb_enabled_var = tk.BooleanVar(value=cfg.get("supabase_enabled", False))
        tk.Checkbutton(
            f, text="Supabaseへの送信を有効にする",
            variable=self._sb_enabled_var, font=("", 10, "bold")
        ).pack(anchor="w", padx=10, pady=(0, 6))

        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=4)

        r1 = tk.Frame(f)
        r1.pack(fill="x", padx=10, pady=3)
        tk.Label(r1, text="Project URL：", width=16, anchor="w").pack(side="left")
        self._sb_url_var = tk.StringVar(value=cfg.get("supabase_url", ""))
        tk.Entry(r1, textvariable=self._sb_url_var, width=40).pack(side="left", padx=4)

        r2 = tk.Frame(f)
        r2.pack(fill="x", padx=10, pady=3)
        tk.Label(r2, text="Service Role Key：", width=16, anchor="w").pack(side="left")
        self._sb_key_var = tk.StringVar(value=cfg.get("supabase_key", ""))
        tk.Entry(r2, textvariable=self._sb_key_var, show="*", width=40).pack(side="left", padx=4)

        btn_row = tk.Frame(f)
        btn_row.pack(fill="x", padx=10, pady=(8, 4))
        tk.Button(
            btn_row, text="設定を保存", width=12, bg="#c8f0d8",
            command=self._save_cloud_config
        ).pack(side="left", padx=4)
        tk.Button(
            btn_row, text="接続テスト", width=12, bg="#d2e6ff",
            command=self._test_supabase_connection
        ).pack(side="left", padx=4)
        self._sb_test_label = tk.Label(btn_row, text="", fg="gray")
        self._sb_test_label.pack(side="left", padx=8)

        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=10)

        tk.Label(f, text="Supabaseセットアップ手順", font=("", 10, "bold"), anchor="w").pack(
            fill="x", padx=10
        )
        steps = (
            "1. supabase.com でプロジェクトを作成\n"
            "2. SQL Editorで下記を実行してテーブルを作成：\n\n"
            "   create table public.utterances (\n"
            "     id bigserial primary key,\n"
            "     session_id text not null,\n"
            "     created_at timestamptz default now(),\n"
            "     ts text, speaker text, text text,\n"
            "     tag text, summary text\n"
            "   );\n"
            "   alter table public.utterances enable row level security;\n"
            "   create policy \"read\" on public.utterances for select using (true);\n"
            "   create policy \"insert\" on public.utterances for insert with check (true);\n\n"
            "3. Project Settings → API から URL と Service Role Key をコピー\n"
            "4. 上のフォームに貼り付けて「設定を保存」"
        )
        txt = tk.Text(f, height=16, wrap="none", bg="#f5f5f5", font=("Courier", 9))
        txt.insert("1.0", steps)
        txt.config(state="disabled")
        txt.pack(fill="x", padx=10, pady=(4, 10))

    def _save_cloud_config(self):
        self._ai_cfg.update({
            "supabase_url": self._sb_url_var.get().strip(),
            "supabase_key": self._sb_key_var.get().strip(),
            "supabase_enabled": self._sb_enabled_var.get(),
            "web_mode": self._web_mode_var.get(),
        })
        config_manager.save_config(self._ai_cfg)
        self._apply_supabase_config()
        messagebox.showinfo("保存完了", "クラウド設定を保存しました")

    def _apply_supabase_config(self):
        cfg = self._ai_cfg
        url = self._sb_url_var.get().strip() if hasattr(self, "_sb_url_var") else cfg.get("supabase_url", "")
        key = self._sb_key_var.get().strip() if hasattr(self, "_sb_key_var") else cfg.get("supabase_key", "")
        enabled = self._sb_enabled_var.get() if hasattr(self, "_sb_enabled_var") else cfg.get("supabase_enabled", False)
        if enabled and url and key:
            supabase_client.init(url, key)
        else:
            supabase_client.init("", "")

    def _test_supabase_connection(self):
        url = self._sb_url_var.get().strip()
        key = self._sb_key_var.get().strip()
        if not url or not key:
            self._sb_test_label.config(text="URL と Key を入力してください", fg="red")
            return
        self._sb_test_label.config(text="テスト中...", fg="gray")
        self.root.update_idletasks()

        def worker():
            try:
                from supabase import create_client
                client = create_client(url, key)
                client.table("utterances").select("id").limit(1).execute()
                msg, color = "接続成功", "green"
            except Exception as e:
                msg, color = f"エラー: {e}", "red"
            self.root.after(0, lambda m=msg, c=color: self._sb_test_label.config(text=m, fg=c))

        threading.Thread(target=worker, daemon=True).start()

    # ─────────────────────────────── Excelビューア（右ペイン）

    def _build_sheet_viewer(self, parent: tk.Frame):
        header_row = tk.Frame(parent)
        header_row.pack(fill="x", padx=6, pady=(6, 2))
        tk.Label(header_row, text="Excelビューア", font=("", 10, "bold")).pack(side="left")
        tk.Button(
            header_row, text="↺ 更新", width=6,
            command=self._refresh_viewer
        ).pack(side="right")
        self._viewer_status = tk.Label(header_row, text="", fg="gray", font=("", 9))
        self._viewer_status.pack(side="right", padx=6)

        sheet_frame = tk.Frame(parent)
        sheet_frame.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        self._viewer_sheet = TkSheet(
            sheet_frame,
            headers=["時刻", "話者", "発言内容", "タグ", "要約"],
            show_x_scrollbar=True,
            show_y_scrollbar=True,
            theme="light green",
        )
        self._viewer_sheet.pack(fill="both", expand=True)

        # コピペ・選択を有効化
        self._viewer_sheet.enable_bindings(
            "single_select",
            "drag_select",
            "column_select",
            "row_select",
            "copy",
            "rc_select",
        )
        self._viewer_sheet.column_width(column=0, width=72)   # 時刻
        self._viewer_sheet.column_width(column=1, width=70)   # 話者
        self._viewer_sheet.column_width(column=2, width=260)  # 発言内容
        self._viewer_sheet.column_width(column=3, width=60)   # タグ
        self._viewer_sheet.column_width(column=4, width=200)  # 要約

    def _refresh_viewer(self):
        path = self._excel_path_var.get() if hasattr(self, "_excel_path_var") else ""
        sheet_name = self._sheet_var.get() if hasattr(self, "_sheet_var") else ""
        if not path or not sheet_name or self._viewer_sheet is None:
            return

        def _load():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return None, "シートが見つかりません"
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue  # ヘッダー行をスキップ
                    if all(v is None for v in row):
                        continue
                    rows.append([("" if v is None else str(v)) for v in row[:5]])
                wb.close()
                return rows, None
            except Exception as e:
                return None, str(e)

        def _worker():
            rows, err = _load()
            self.root.after(0, lambda: self._apply_viewer_data(rows, err))

        threading.Thread(target=_worker, daemon=True).start()

    def _apply_viewer_data(self, rows: list | None, err: str | None):
        if self._viewer_sheet is None:
            return
        if err:
            self._viewer_status.config(text=f"エラー: {err}", fg="red")
            return
        if rows is None:
            return

        self._viewer_sheet.set_sheet_data(rows)

        # 最終行を自動スクロール（新規行追加時のみ）
        if len(rows) > self._viewer_row_count and rows:
            self._viewer_sheet.see(row=len(rows) - 1, column=0)
        self._viewer_row_count = len(rows)

        now = datetime.now().strftime("%H:%M:%S")
        self._viewer_status.config(text=f"更新 {now}　{len(rows)}行", fg="gray")

    # ─────────────────────────────── AI設定タブ

    def _build_ai_tab(self):
        f = self._tab_ai
        cfg = self._ai_cfg

        # 有効/無効トグル
        top_row = tk.Frame(f)
        top_row.pack(fill="x", padx=10, pady=(10, 4))
        self._ai_enabled_var = tk.BooleanVar(value=cfg.get("ai_enabled", False))
        tk.Checkbutton(
            top_row, text="AI要約・タグ付けを有効にする",
            variable=self._ai_enabled_var, font=("", 10, "bold"),
            command=self._on_ai_toggle
        ).pack(side="left")

        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=4)

        # プロバイダー
        r1 = tk.Frame(f)
        r1.pack(fill="x", padx=10, pady=2)
        tk.Label(r1, text="プロバイダー：", width=16, anchor="w").pack(side="left")
        self._ai_provider_var = tk.StringVar(value=cfg.get("provider", "ChatGPT"))
        providers = list(ai_module.PROVIDERS.keys())
        ttk.Combobox(
            r1, textvariable=self._ai_provider_var,
            values=providers, state="readonly", width=20
        ).pack(side="left", padx=4)
        self._ai_provider_var.trace_add("write", self._on_provider_change)

        # モデル
        r2 = tk.Frame(f)
        r2.pack(fill="x", padx=10, pady=2)
        tk.Label(r2, text="モデル：", width=16, anchor="w").pack(side="left")
        self._ai_model_var = tk.StringVar(value=cfg.get("model", "gpt-4o-mini"))
        self._ai_model_cb = ttk.Combobox(
            r2, textvariable=self._ai_model_var,
            state="readonly", width=28
        )
        self._ai_model_cb.pack(side="left", padx=4)
        self._refresh_model_list()

        # APIキー
        r3 = tk.Frame(f)
        r3.pack(fill="x", padx=10, pady=2)
        tk.Label(r3, text="APIキー：", width=16, anchor="w").pack(side="left")
        self._ai_key_var = tk.StringVar(value=cfg.get("api_key", ""))
        tk.Entry(r3, textvariable=self._ai_key_var, show="*", width=38).pack(side="left", padx=4)

        # Azureエンドポイント（Azure/Copilot選択時のみ表示）
        self._azure_row = tk.Frame(f)
        tk.Label(self._azure_row, text="Azureエンドポイント：", width=20, anchor="w").pack(side="left")
        self._ai_endpoint_var = tk.StringVar(value=cfg.get("azure_endpoint", ""))
        tk.Entry(self._azure_row, textvariable=self._ai_endpoint_var, width=34).pack(side="left", padx=4)
        self._ai_key_row = r3  # Azureエンドポイント行の挿入位置の基準

        # 予算・コンテキスト
        r5 = tk.Frame(f)
        r5.pack(fill="x", padx=10, pady=2)
        tk.Label(r5, text="予算上限（円）：", width=16, anchor="w").pack(side="left")
        self._ai_budget_var = tk.StringVar(value=str(cfg.get("budget_jpy", 100.0)))
        tk.Spinbox(r5, textvariable=self._ai_budget_var, from_=1, to=99999, increment=10, width=8).pack(side="left", padx=4)
        tk.Label(r5, text="　コンテキスト件数：").pack(side="left")
        self._ai_ctx_var = tk.StringVar(value=str(cfg.get("context_window", 10)))
        tk.Spinbox(r5, textvariable=self._ai_ctx_var, from_=1, to=30, increment=1, width=4).pack(side="left", padx=4)

        # 保存・テストボタン
        btn_row = tk.Frame(f)
        btn_row.pack(fill="x", padx=10, pady=(6, 4))
        tk.Button(btn_row, text="設定を保存", width=12, bg="#c8f0d8", command=self._save_ai_config).pack(side="left", padx=4)
        tk.Button(btn_row, text="接続テスト", width=12, bg="#d2e6ff", command=self._test_ai_connection).pack(side="left", padx=4)
        self._ai_test_label = tk.Label(btn_row, text="", fg="gray")
        self._ai_test_label.pack(side="left", padx=8)

        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=4)

        # 円グラフ + クレジット情報
        credit_frame = tk.Frame(f)
        credit_frame.pack(fill="x", padx=10, pady=4)

        self._pie_canvas = tk.Canvas(credit_frame, width=120, height=120, bg="white", highlightthickness=0)
        self._pie_canvas.pack(side="left", padx=(0, 12))

        info_frame = tk.Frame(credit_frame)
        info_frame.pack(side="left", fill="y")
        tk.Label(info_frame, text="クレジット使用状況", font=("", 10, "bold"), anchor="w").pack(anchor="w")
        self._credit_used_label = tk.Label(info_frame, text="使用：¥0.000", anchor="w")
        self._credit_used_label.pack(anchor="w")
        self._credit_remain_label = tk.Label(info_frame, text="残り：¥---", anchor="w")
        self._credit_remain_label.pack(anchor="w")
        self._credit_pct_label = tk.Label(info_frame, text="残り：---%", fg="gray", anchor="w")
        self._credit_pct_label.pack(anchor="w")

        # バッチ切替ボタン
        self._batch_btn = tk.Button(
            info_frame, text="まとめて処理する方式に切替",
            bg="#ffe0b2", width=22, command=self._switch_to_batch
        )
        self._batch_btn.pack(anchor="w", pady=(8, 0))
        self._batch_run_btn = tk.Button(
            info_frame, text="未処理発話をまとめてAI処理",
            bg="#c8f0d8", width=22, state="disabled", command=self._run_batch_ai
        )
        self._batch_run_btn.pack(anchor="w", pady=(4, 0))

        self._draw_pie(0.0)
        self._update_azure_row_visibility()

    def _on_provider_change(self, *_):
        self._refresh_model_list()
        self._update_azure_row_visibility()

    def _refresh_model_list(self):
        provider = self._ai_provider_var.get()
        models = ai_module.PROVIDERS.get(provider, {}).get("models", [])
        self._ai_model_cb.config(values=models)
        current = self._ai_model_var.get()
        if current not in models and models:
            self._ai_model_var.set(models[0])

    def _update_azure_row_visibility(self):
        if self._ai_provider_var.get() == "Azure/Copilot":
            self._azure_row.pack(fill="x", padx=10, pady=2, after=self._ai_key_row)
        else:
            self._azure_row.pack_forget()

    def _on_ai_toggle(self):
        self._apply_ai_config()

    def _save_ai_config(self):
        try:
            budget = float(self._ai_budget_var.get())
            ctx = int(self._ai_ctx_var.get())
        except ValueError:
            messagebox.showwarning("入力エラー", "予算・コンテキスト件数は数値で入力してください")
            return
        self._ai_cfg.update({
            "provider": self._ai_provider_var.get(),
            "model": self._ai_model_var.get(),
            "api_key": self._ai_key_var.get(),
            "azure_endpoint": self._ai_endpoint_var.get(),
            "budget_jpy": budget,
            "context_window": ctx,
            "ai_enabled": self._ai_enabled_var.get(),
        })
        config_manager.save_config(self._ai_cfg)
        self._apply_ai_config()
        messagebox.showinfo("保存完了", "AI設定を保存しました")

    def _apply_ai_config(self):
        cfg = self._ai_cfg
        enabled = self._ai_enabled_var.get() if hasattr(self, "_ai_enabled_var") else cfg.get("ai_enabled", False)
        key = self._ai_key_var.get() if hasattr(self, "_ai_key_var") else cfg.get("api_key", "")
        provider = self._ai_provider_var.get() if hasattr(self, "_ai_provider_var") else cfg.get("provider", "ChatGPT")
        model = self._ai_model_var.get() if hasattr(self, "_ai_model_var") else cfg.get("model", "gpt-4o-mini")
        endpoint = self._ai_endpoint_var.get() if hasattr(self, "_ai_endpoint_var") else cfg.get("azure_endpoint", "")
        if enabled and key:
            self._ai_client = ai_module.AIClient(
                provider=provider, api_key=key, model=model, endpoint=endpoint
            )
        else:
            self._ai_client = None

    def _test_ai_connection(self):
        key = self._ai_key_var.get().strip()
        if not key:
            self._ai_test_label.config(text="APIキーを入力してください", fg="red")
            return
        test_client = ai_module.AIClient(
            provider=self._ai_provider_var.get(),
            api_key=key,
            model=self._ai_model_var.get(),
            endpoint=self._ai_endpoint_var.get(),
        )
        self._ai_test_label.config(text="テスト中...", fg="gray")
        self.root.update_idletasks()

        def worker():
            try:
                ok, msg = test_client.test_connection()
            except Exception as e:
                ok, msg = False, f"予期しないエラー: {e}"
            color = "green" if ok else "red"
            self.root.after(0, lambda m=msg, c=color: self._ai_test_label.config(text=m, fg=c))

        threading.Thread(target=worker, daemon=True).start()

    # ─────────────────────────────── クレジット管理

    def _update_credit_usage(self, input_tokens: int, output_tokens: int):
        if not self._ai_client:
            return
        cost = self._ai_client.estimate_cost_jpy(input_tokens, output_tokens)
        self._session_cost_jpy += cost

        try:
            budget = float(self._ai_budget_var.get())
        except Exception:
            budget = self._ai_cfg.get("budget_jpy", 100.0)

        used = self._session_cost_jpy
        remain = max(0.0, budget - used)
        pct = remain / budget * 100 if budget > 0 else 0.0

        self._credit_used_label.config(text=f"使用：¥{used:.3f}")
        self._credit_remain_label.config(text=f"残り：¥{remain:.3f}")
        self._credit_pct_label.config(text=f"残り：{pct:.1f}%")
        self._draw_pie(used / budget if budget > 0 else 1.0)

        if pct <= 10.0 and not self._budget_warned:
            self._budget_warned = True
            self.root.after(0, self._warn_low_budget, pct)

        if used >= budget and not self._budget_exhausted:
            self._budget_exhausted = True
            self.root.after(0, lambda: messagebox.showinfo(
                "クレジット枯渇", "予算上限に達したためAI要約・タグ付けを停止しました。"
            ))

    def _warn_low_budget(self, pct: float):
        ans = messagebox.askyesno(
            "クレジット残量低下",
            f"残り {pct:.1f}% です。\n"
            "まとめて処理する方式に切り替えますか？\n"
            "（「はい」で発話ごとの自動処理を停止し、手動実行に切り替えます）"
        )
        if ans:
            self._switch_to_batch()

    def _switch_to_batch(self):
        self._batch_mode = True
        if hasattr(self, "_batch_btn"):
            self._batch_btn.config(state="disabled", text="まとめて処理（切替済）")
        if hasattr(self, "_batch_run_btn"):
            self._batch_run_btn.config(state="normal")

    def _draw_pie(self, used_ratio: float):
        if not hasattr(self, "_pie_canvas"):
            return
        c = self._pie_canvas
        c.delete("all")
        w, h = 120, 120
        margin = 10
        x0, y0, x1, y1 = margin, margin, w - margin, h - margin

        used_ratio = max(0.0, min(1.0, used_ratio))
        remain_ratio = 1.0 - used_ratio

        if used_ratio <= 0.0:
            c.create_oval(x0, y0, x1, y1, fill="#d0e8d0", outline="")
        elif used_ratio >= 1.0:
            c.create_oval(x0, y0, x1, y1, fill="#e05050", outline="")
        else:
            used_deg = used_ratio * 360.0
            if used_ratio < 0.5:
                used_color = "#4caf50"
            elif used_ratio < 0.9:
                used_color = "#ff9800"
            else:
                used_color = "#e05050"
            c.create_arc(x0, y0, x1, y1, start=90, extent=-used_deg, fill=used_color, outline="")
            c.create_arc(x0, y0, x1, y1, start=90 - used_deg, extent=-(360 - used_deg), fill="#cccccc", outline="")

        pct = (1.0 - used_ratio) * 100
        pct_color = "#e05050" if pct <= 10 else "#333333"
        c.create_text(w // 2, h // 2, text=f"残り\n{pct:.0f}%", font=("", 10, "bold"), fill=pct_color)

        if pct <= 10:
            c.create_oval(x0 - 2, y0 - 2, x1 + 2, y1 + 2, outline="#e05050", width=3)

    def _build_rec_tab(self):
        f = self._tab_rec

        # ファイル選択
        row = tk.Frame(f)
        row.pack(fill="x", padx=10, pady=(10, 2))
        tk.Label(row, text="出力先：").pack(side="left")
        self._excel_path_var = tk.StringVar()
        tk.Entry(row, textvariable=self._excel_path_var, width=32).pack(side="left", padx=6)
        tk.Button(row, text="参照...", command=self._pick_excel).pack(side="left")

        # シート選択
        sheet_row = tk.Frame(f)
        sheet_row.pack(fill="x", padx=10, pady=(0, 4))
        tk.Label(sheet_row, text="シート：").pack(side="left")
        self._sheet_var = tk.StringVar()
        self._sheet_cb = ttk.Combobox(
            sheet_row, textvariable=self._sheet_var,
            state="disabled", width=28
        )
        self._sheet_cb.pack(side="left", padx=6)
        self._sheet_cb.bind("<<ComboboxSelected>>", self._on_sheet_change)
        tk.Button(sheet_row, text="更新", command=self._refresh_sheets).pack(side="left")

        # セッションID
        sid_row = tk.Frame(f)
        sid_row.pack(fill="x", padx=10, pady=(0, 4))
        tk.Label(sid_row, text="セッションID：").pack(side="left")
        self._session_id_var = tk.StringVar()
        tk.Entry(sid_row, textvariable=self._session_id_var, width=12,
                 font=("Courier", 12)).pack(side="left", padx=4)
        tk.Label(sid_row, text="（空白で自動生成）", fg="gray",
                 font=("", 9)).pack(side="left")

        # 検出設定フレーム
        det_frame = ttk.LabelFrame(f, text="検出設定")
        det_frame.pack(fill="x", padx=10, pady=(4, 4))

        # 検出モード
        mode_row = tk.Frame(det_frame)
        mode_row.pack(fill="x", padx=6, pady=(4, 2))
        tk.Label(mode_row, text="検出方式：").pack(side="left")
        tk.Radiobutton(
            mode_row, text="PTT信号（有線接続）",
            variable=self._detect_mode, value="ptt",
            command=self._on_mode_change
        ).pack(side="left", padx=4)
        tk.Radiobutton(
            mode_row, text="VAD（マイク/デバッグ）",
            variable=self._detect_mode, value="vad",
            command=self._on_mode_change
        ).pack(side="left", padx=4)

        # PTT閾値 + レベルメーター
        thr_row = tk.Frame(det_frame)
        thr_row.pack(fill="x", padx=6, pady=(2, 6))
        tk.Label(thr_row, text="PTT閾値：").pack(side="left")
        self._ptt_thr_entry = tk.Entry(
            thr_row, textvariable=self._ptt_threshold_var, width=7
        )
        self._ptt_thr_entry.pack(side="left", padx=(0, 10))

        tk.Label(thr_row, text="信号レベル：").pack(side="left")
        self._level_label = tk.Label(
            thr_row, text="─" * METER_BARS,
            font=("Courier", 11), fg="gray", width=METER_BARS + 2
        )
        self._level_label.pack(side="left")
        self._ptt_state_label = tk.Label(thr_row, text="", width=10)
        self._ptt_state_label.pack(side="left", padx=4)

        # 開始・停止ボタン
        row2 = tk.Frame(f)
        row2.pack(fill="x", padx=10, pady=4)
        self._btn_rec_start = tk.Button(
            row2, text="記録開始", width=10, bg="#aadcb4",
            font=("", 10, "bold"), command=self._rec_start
        )
        self._btn_rec_start.pack(side="left", padx=4)
        self._btn_rec_stop = tk.Button(
            row2, text="記録停止", width=10, bg="#e1e1e1",
            state="disabled", command=self._rec_stop
        )
        self._btn_rec_stop.pack(side="left", padx=4)

        # ステータスバー
        self._status_var = tk.StringVar(value="● 待機中")
        tk.Label(
            f, textvariable=self._status_var,
            bg="#d2e6fa", anchor="w", padx=10, pady=4
        ).pack(fill="x")

        # 直前の記録
        tk.Label(f, text="直前の記録（修正可）：", anchor="w").pack(fill="x", padx=10, pady=(6, 2))
        self._edit_text = tk.Text(f, height=3, wrap="word")
        self._edit_text.pack(fill="x", padx=10)

        fix_row = tk.Frame(f)
        fix_row.pack(fill="x", padx=10, pady=(2, 4))
        self._btn_fix = tk.Button(
            fix_row, text="直前を修正・再出力", bg="#ffe0b2",
            state="disabled", command=self._fix_last
        )
        self._btn_fix.pack(side="right")

        # ログ
        ttk.Separator(f, orient="horizontal").pack(fill="x", padx=6, pady=4)
        tk.Label(f, text="出力済みログ：", anchor="w").pack(fill="x", padx=10)

        log_frame = tk.Frame(f)
        log_frame.pack(fill="both", expand=True, padx=10, pady=(2, 10))
        log_scroll = tk.Scrollbar(log_frame, orient="vertical")
        self._log_text = tk.Text(
            log_frame, height=6, state="disabled",
            wrap="word", bg="#f8f8f8",
            yscrollcommand=log_scroll.set
        )
        log_scroll.config(command=self._log_text.yview)
        self._log_text.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")

    def _on_mode_change(self):
        mode = self._detect_mode.get()
        state = "normal" if mode == "ptt" else "disabled"
        self._ptt_thr_entry.config(state=state)

    # ─────────────────────────────── DBタブ操作

    def _db_start_recording(self):
        name = self._db_name_var.get().strip()
        if not name:
            messagebox.showwarning("入力エラー", "名前を入力してください")
            return

        self._db_audio_buf = []
        self.is_db_recording = True
        self._db_capturer = ac.AudioCapture(device=self._get_selected_device())
        self._db_capturer.start()
        self._btn_db_start.config(state="disabled")
        self._btn_db_stop.config(state="normal")
        threading.Thread(target=self._db_accumulate, daemon=True).start()

    def _db_accumulate(self):
        import time
        while self.is_db_recording:
            chunk = self._db_capturer.get_chunk()
            if chunk is not None:
                self._db_audio_buf.append(chunk)
            else:
                time.sleep(0.005)

    def _db_stop_recording(self):
        self.is_db_recording = False
        if self._db_capturer:
            self._db_capturer.stop()
        self._btn_db_start.config(state="normal")
        self._btn_db_stop.config(state="disabled")

        if not self._db_audio_buf:
            return
        audio = np.concatenate(self._db_audio_buf)
        duration = len(audio) / SAMPLE_RATE
        name = self._db_name_var.get().strip()
        label = f"{name}：{duration:.1f}秒"
        self._current_samples.append((audio, label))
        self._sample_lb.insert("end", label)

    def _db_delete_sample(self):
        sel = self._sample_lb.curselection()
        if not sel:
            return
        del self._current_samples[sel[0]]
        self._sample_lb.delete(0, "end")
        for _, label in self._current_samples:
            self._sample_lb.insert("end", label)

    def _db_save(self):
        name = self._db_name_var.get().strip()
        if not name:
            messagebox.showwarning("入力エラー", "名前を入力してください")
            return
        if not self._current_samples:
            messagebox.showwarning("サンプルなし", "録音サンプルがありません")
            return

        def worker():
            try:
                embeddings = [speaker_id.get_embedding(a) for a, _ in self._current_samples]
                db_manager.add_speaker(name, embeddings)
                self.root.after(0, lambda: messagebox.showinfo("保存完了", f"「{name}」の声紋を保存しました"))
                self.root.after(0, self._db_reset)
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("エラー", str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _db_reset(self):
        self._current_samples = []
        self._sample_lb.delete(0, "end")
        self._db_name_var.set("")

    # ─────────────────────────────── 記録タブ操作

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel ファイル", "*.xlsx *.xls"), ("すべて", "*.*")]
        )
        if path:
            self._excel_path_var.set(path)
            excel_writer.set_target(path)
            self._known_sheets = []  # キャッシュリセットして強制更新
            self._refresh_sheets()

    def _refresh_sheets(self, auto: bool = False) -> None:
        """シート名一覧を取得して Combobox を更新する"""
        path = self._excel_path_var.get()
        if not path:
            return
        try:
            sheets = excel_writer.get_sheets(path)
        except Exception:
            return

        if auto and sheets == self._known_sheets:
            return  # 変化なし

        self._known_sheets = sheets
        current = self._sheet_var.get()
        self._sheet_cb.config(values=sheets, state="readonly")

        if current in sheets:
            # 現在の選択を維持
            self._sheet_var.set(current)
        else:
            # 先頭シートを自動選択
            if sheets:
                self._sheet_var.set(sheets[0])
                excel_writer.set_sheet(sheets[0])
            else:
                self._sheet_var.set("")
                excel_writer.set_sheet(None)

    def _on_sheet_change(self, event=None) -> None:
        name = self._sheet_var.get()
        excel_writer.set_sheet(name if name else None)

    def _rec_start(self):
        if not self._excel_path_var.get():
            messagebox.showwarning("未設定", "出力先Excelファイルを指定してください")
            return

        self.is_recording = True
        manual_id = self._session_id_var.get().strip().upper()
        if manual_id:
            self._session_id = manual_id
        else:
            import uuid
            self._session_id = str(uuid.uuid4())[:8].upper()
        self._ai_context.clear()
        self._session_cost_jpy = 0.0
        self._budget_exhausted = False
        self._budget_warned = False
        self._batch_pending.clear()
        if hasattr(self, "_pie_canvas"):
            self._draw_pie(0.0)
        if hasattr(self, "_credit_used_label"):
            self._credit_used_label.config(text="使用：¥0.000")
            self._credit_remain_label.config(text=f"残り：¥{self._ai_cfg.get('budget_jpy', 100.0):.3f}")
            self._credit_pct_label.config(text="残り：100.0%")

        self._capturer = ac.AudioCapture(device=self._get_selected_device())

        mode = self._detect_mode.get()
        if mode == "ptt":
            try:
                thr = float(self._ptt_threshold_var.get())
            except ValueError:
                thr = 0.002
            self._detector = vad_module.PTTDetector(energy_threshold=thr)
        else:
            self._detector = vad_module.VAD()

        self._capturer.start()
        self._btn_rec_start.config(state="disabled")
        self._btn_rec_stop.config(state="normal")
        self._status_var.set("● 待受中　─　発言を検出したら自動で記録します")

        self._vad_thread = threading.Thread(target=self._detect_loop, daemon=True)
        self._vad_thread.start()

    def _rec_stop(self):
        self.is_recording = False
        if self._capturer:
            self._capturer.stop()
        if self._detector:
            self._detector.reset()

        self._btn_rec_start.config(state="normal")
        self._btn_rec_stop.config(state="disabled")
        self._status_var.set("● 待機中")
        self._update_level_meter(0.0, ptt_active=False)

    def _detect_loop(self):
        import time
        while self.is_recording:
            chunk = self._capturer.get_chunk()
            if chunk is not None:
                segment, level = self._detector.process(chunk)
                ptt = self._detector.is_ptt_active
                # レベルメーター更新はUIスレッドへ（100msに1回で十分）
                self._result_queue.put({"type": "level", "level": level, "ptt": ptt})
                if segment is not None:
                    self._utterance_queue.put(segment.copy())
            else:
                time.sleep(0.005)

    # ─────────────────────────────── 処理ワーカー（常時稼働・並列処理）

    def _start_process_worker(self):
        threading.Thread(target=self._process_worker, daemon=True).start()

    def _process_worker(self):
        while True:
            audio = self._utterance_queue.get()
            if audio is None:
                break
            try:
                self._result_queue.put({"type": "status", "text": "⏳ 解析中..."})

                # ── Webモード：話者識別＋音声アップロードのみ
                if self._web_mode_var.get():
                    spk, sim = speaker_id.identify(audio)
                    ts = datetime.now().strftime("%H:%M:%S")
                    if supabase_client.is_enabled():
                        supabase_client.upload_audio_and_insert_pending(
                            self._session_id, ts, spk, audio
                        )
                        self._result_queue.put({
                            "type": "web_uploaded",
                            "speaker": spk,
                            "similarity": int(sim * 100),
                            "timestamp": ts,
                        })
                    else:
                        self._result_queue.put({
                            "type": "status",
                            "text": "⚠ Supabase未設定 — クラウド設定タブを確認してください",
                        })
                    continue

                # ── ローカルモード：speaker IDとWhisperを並列実行
                import transcriber  # オフラインモード時のみ使用（遅延import）
                with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
                    f_spk = ex.submit(speaker_id.identify, audio)
                    f_txt = ex.submit(transcriber.transcribe, audio)
                    spk, sim = f_spk.result()
                    text = f_txt.result()

                # 空文字はハルシネーション or 無音と判定されたため書き込まない
                if not text:
                    self._result_queue.put({"type": "status", "text": "（無音/ノイズのためスキップ）"})
                    continue

                ts = datetime.now().strftime("%H:%M:%S")

                # AI要約・タグ付け
                ai_tag, ai_summary = "", ""
                if (
                    self._ai_client is not None
                    and not self._budget_exhausted
                    and not self._batch_mode
                ):
                    try:
                        ctx_n = int(self._ai_cfg.get("context_window", 10))
                        ai_result = self._ai_client.summarize_and_tag(
                            text, self._ai_context[-ctx_n:]
                        )
                        ai_tag = ai_result.get("tag", "")
                        ai_summary = ai_result.get("summary", "")
                        self.root.after(0, self._update_credit_usage,
                                        ai_result.get("input_tokens", 0),
                                        ai_result.get("output_tokens", 0))
                    except Exception as e:
                        ai_summary = f"[AIエラー: {e}]"

                self._result_queue.put({
                    "type": "result",
                    "speaker": spk,
                    "similarity": sim,
                    "text": text,
                    "timestamp": ts,
                    "tag": ai_tag,
                    "summary": ai_summary,
                })
            except Exception as e:
                self._result_queue.put({"type": "error", "message": str(e)})

    # ─────────────────────────────── キューポーリング

    def _poll_queue(self):
        # レベルメーターは最新値のみ使う（連続して積まれても最後だけ反映）
        latest_level = None
        try:
            while True:
                item = self._result_queue.get_nowait()
                if item["type"] == "level":
                    latest_level = item
                elif item["type"] == "status":
                    self._status_var.set(f"● 待受中　{item['text']}")
                elif item["type"] == "result":
                    self._auto_output(item)
                elif item["type"] == "web_uploaded":
                    spk = item["speaker"]
                    ts = item["timestamp"]
                    sim = item["similarity"]
                    self._status_var.set(
                        f"● 待受中　{ts}　{spk}（{sim}%）→ ☁ Webに送信済"
                    )
                    entry = f"{ts}　{spk}　☁ Web処理中...\n"
                    self._log_text.config(state="normal")
                    self._log_text.insert("end", entry)
                    self._log_text.see("end")
                    self._log_text.config(state="disabled")
                elif item["type"] == "error":
                    self._status_var.set(f"エラー: {item['message']}")
        except queue.Empty:
            pass

        if latest_level is not None:
            self._update_level_meter(latest_level["level"], latest_level["ptt"])

        now = time.monotonic()

        # シート一覧を3秒ごとに自動更新
        if self._excel_path_var.get() and now - self._last_sheet_check >= 3.0:
            self._last_sheet_check = now
            self._refresh_sheets(auto=True)

        # ビューアを3秒ごとに自動更新（記録中は毎回、停止中は変化があれば）
        if self._excel_path_var.get() and now - self._last_viewer_refresh >= 3.0:
            self._last_viewer_refresh = now
            self._refresh_viewer()

        self.root.after(80, self._poll_queue)

    def _update_level_meter(self, rms: float, ptt_active: bool):
        try:
            thr = float(self._ptt_threshold_var.get())
        except ValueError:
            thr = 0.002

        # 0〜0.1 を METER_BARS本のバーにマッピング
        filled = min(METER_BARS, int(rms / 0.1 * METER_BARS))
        bar = "█" * filled + "─" * (METER_BARS - filled)

        if ptt_active:
            self._level_label.config(text=bar, fg="red")
            self._ptt_state_label.config(text="🔴 送信中", fg="red")
        elif rms > thr:
            self._level_label.config(text=bar, fg="orange")
            self._ptt_state_label.config(text="", fg="black")
        else:
            self._level_label.config(text=bar, fg="gray")
            self._ptt_state_label.config(text="", fg="black")

    # ─────────────────────────────── Excel出力

    def _auto_output(self, item: dict):
        spk = item["speaker"]
        text = item["text"]
        ts = item["timestamp"]
        sim_pct = int(item["similarity"] * 100)
        tag = item.get("tag", "")
        summary = item.get("summary", "")

        try:
            excel_writer.append_row(speaker=spk, text=text, tag=tag, summary=summary)
        except Exception as e:
            self._status_var.set(f"Excel書き込みエラー: {e}")
            return

        # Supabaseに非同期でinsert
        if self._session_id:
            threading.Thread(
                target=supabase_client.insert_utterance,
                args=(self._session_id, ts, spk, text, tag, summary),
                daemon=True,
            ).start()

        # バッチモード用に未処理として保持
        if self._batch_mode and self._ai_client and not self._budget_exhausted:
            self._batch_pending.append({"speaker": spk, "text": text, "tag": "", "summary": ""})

        # コンテキストに追加
        self._ai_context.append({"speaker": spk, "text": text, "tag": tag, "summary": summary})

        # 書き込み直後にビューアを即時更新
        self._last_viewer_refresh = 0.0

        self._status_var.set(
            f"● 待受中　{ts}　{spk}（{sim_pct}%）→ 書き込み済"
            + (f"　タグ:{tag}" if tag else "")
        )

        self._edit_text.delete("1.0", "end")
        self._edit_text.insert("1.0", text)
        self._pending_speaker = spk
        self._pending_timestamp = ts
        self._btn_fix.config(state="normal")

        entry = f"{ts}　{spk}　{text}\n"
        self._log_text.config(state="normal")
        self._log_text.insert("end", entry)
        self._log_text.see("end")
        self._log_text.config(state="disabled")

    def _fix_last(self):
        text = self._edit_text.get("1.0", "end").strip()
        if not text:
            return
        try:
            excel_writer.update_last_row(speaker=self._pending_speaker, text=text)
        except Exception as e:
            messagebox.showerror("修正エラー", str(e))
            return

        self._log_text.config(state="normal")
        content = self._log_text.get("1.0", "end")
        lines = [l for l in content.split("\n") if l.strip()]
        if lines:
            lines[-1] = f"{self._pending_timestamp}　{self._pending_speaker}　{text}（修正済）"
        self._log_text.delete("1.0", "end")
        self._log_text.insert("1.0", "\n".join(lines) + "\n")
        self._log_text.see("end")
        self._log_text.config(state="disabled")
        self._btn_fix.config(state="disabled")

    # ─────────────────────────────── バッチAI処理

    def _run_batch_ai(self):
        if not self._ai_client:
            messagebox.showwarning("AI未設定", "AI設定タブでAPIキーを設定してください")
            return
        if not self._batch_pending:
            messagebox.showinfo("対象なし", "未処理の発話がありません")
            return

        pending = list(self._batch_pending)
        self._batch_pending.clear()
        self._batch_run_btn.config(state="disabled", text="処理中...")

        def worker():
            context: list[dict] = []
            ctx_n = int(self._ai_cfg.get("context_window", 10))
            for item in pending:
                if self._budget_exhausted:
                    break
                try:
                    result = self._ai_client.summarize_and_tag(item["text"], context[-ctx_n:])
                    item["tag"] = result.get("tag", "")
                    item["summary"] = result.get("summary", "")
                    self.root.after(0, self._update_credit_usage,
                                    result.get("input_tokens", 0),
                                    result.get("output_tokens", 0))
                    context.append(item)
                except Exception as e:
                    item["summary"] = f"[AIエラー: {e}]"
            self.root.after(0, self._on_batch_done, pending)

        threading.Thread(target=worker, daemon=True).start()

    def _on_batch_done(self, processed: list[dict]):
        self._batch_run_btn.config(state="normal", text="未処理発話をまとめてAI処理")
        messagebox.showinfo("バッチ処理完了",
            f"{len(processed)}件の発話を処理しました。\n"
            "※Excelへの書き戻しは次バージョンで対応予定。"
        )
